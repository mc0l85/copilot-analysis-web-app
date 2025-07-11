
#!/usr/bin/env python3
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import os
import sys
import json
import argparse
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
import warnings
warnings.filterwarnings('ignore')
st.set_page_config(layout="wide")
class CopilotAnalyzer:
    def __init__(self):
        self.target_user_data = None
        self.full_usage_data = None
        self.utilized_metrics_df = None
        self.output_folder_path = None
        
    def log(self, message):
        print(f"[LOG] {message}")
        
    def load_target_users(self, filepath):
        """Load target users file"""
        try:
            self.target_user_data = pd.read_csv(filepath)
            required_cols = ['UserPrincipalName', 'Company', 'Department', 'City', 'ManagerLine']
            
            if not all(col in self.target_user_data.columns for col in required_cols):
                missing_cols = [col for col in required_cols if col not in self.target_user_data.columns]
                raise ValueError(f"Target user file missing required columns: {missing_cols}")
                
            self.log(f"Loaded {len(self.target_user_data)} target users")
            return True
        except Exception as e:
            self.log(f"Error loading target users: {e}")
            return False
            
    def load_usage_reports(self, filepaths):
        """Load usage report files"""
        try:
            all_reports = []
            for file in filepaths:
                try:
                    if file.lower().endswith('.csv'):
                        df = pd.read_csv(file)
                    else:
                        df = pd.read_excel(file)
                    all_reports.append(df)
                    self.log(f"Loaded usage report: {os.path.basename(file)}")
                except Exception as e:
                    self.log(f"Could not read file: {os.path.basename(file)}. Error: {e}")
                    continue
                    
            if not all_reports:
                raise ValueError("No usage reports could be read")
                
            usage_df = pd.concat(all_reports, ignore_index=True)
            usage_df['User Principal Name'] = usage_df['User Principal Name'].str.lower()
            
            # Handle date columns
            date_cols = [col for col in usage_df.columns if 'date' in col.lower()]
            for col in date_cols:
                usage_df[col] = pd.to_datetime(usage_df[col], errors='coerce', format='mixed')
                
            self.full_usage_data = usage_df.copy()
            self.log(f"Loaded {len(usage_df)} usage records")
            return True
        except Exception as e:
            self.log(f"Error loading usage reports: {e}")
            return False
            
    def apply_filters(self, filters):
        """Apply filters to target user data"""
        if self.target_user_data is None:
            return None
            
        filtered_df = self.target_user_data.copy()
        
        # Apply Company Filter
        if filters.get('companies'):
            filtered_df = filtered_df[filtered_df['Company'].isin(filters['companies'])]
            self.log(f"Applied company filter: {len(filtered_df)} users remaining")
            
        # Apply Department Filter
        if filters.get('departments'):
            filtered_df = filtered_df[filtered_df['Department'].isin(filters['departments'])]
            self.log(f"Applied department filter: {len(filtered_df)} users remaining")
            
        # Apply Location Filter
        if filters.get('locations'):
            filtered_df = filtered_df[filtered_df['City'].isin(filters['locations'])]
            self.log(f"Applied location filter: {len(filtered_df)} users remaining")
            
        # Apply Manager Filter
        if filters.get('managers'):
            manager_mask = pd.Series([False] * len(filtered_df), index=filtered_df.index)
            for manager in filters['managers']:
                manager_mask |= filtered_df['ManagerLine'].str.contains(f'\\b{manager}\\b', regex=True, na=False)
            filtered_df = filtered_df[manager_mask]
            self.log(f"Applied manager filter: {len(filtered_df)} users remaining")
            
        return filtered_df
        
    def analyze_users(self, filtered_target_df=None):
        """Perform user analysis"""
        if self.full_usage_data is None:
            raise ValueError("No usage data loaded")
            
        usage_df = self.full_usage_data
        
        # Determine users to analyze
        if filtered_target_df is not None:
            target_users_emails = set(filtered_target_df['UserPrincipalName'].str.lower())
            all_report_emails = set(usage_df['User Principal Name'].unique())
            utilized_emails = target_users_emails.intersection(all_report_emails)
            self.log(f"Analyzing {len(utilized_emails)} of {len(target_users_emails)} target users found in reports")
        else:
            self.log("No target user file provided. Analyzing all users from reports")
            utilized_emails = set(usage_df['User Principal Name'].unique())
            
        if not utilized_emails:
            raise ValueError("No matching users found to analyze")
            
        matched_users_df = usage_df[usage_df['User Principal Name'].isin(utilized_emails)].copy()
        copilot_tool_cols = [col for col in matched_users_df.columns if 'Last activity date of' in col]
        
        # Calculate analysis period
        min_report_date = usage_df['Report Refresh Date'].min()
        max_report_date = usage_df['Report Refresh Date'].max()
        total_months_in_period = (max_report_date.year - min_report_date.year) * 12 + max_report_date.month - min_report_date.month + 1
        
        # Analyze each user
        user_metrics = []
        for email in utilized_emails:
            user_data = matched_users_df[matched_users_df['User Principal Name'] == email]
            
            activity_dates = user_data[copilot_tool_cols].stack().dropna().unique()
            activity_dates = pd.to_datetime(activity_dates)
            
            if len(activity_dates) == 0:
                first_activity, last_activity = pd.NaT, pd.NaT
                active_months, complexity, avg_tools_per_month, trend = 0, 0, 0, "N/A"
                report_dates = sorted(user_data['Report Refresh Date'].unique())
                first_activity = report_dates[0] if report_dates else pd.NaT
            else:
                first_activity, last_activity = activity_dates.min(), activity_dates.max()
                active_months = len(pd.to_datetime(activity_dates).to_period('M').unique())
                complexity = user_data[copilot_tool_cols].notna().any().sum()
                
                monthly_activity = user_data[copilot_tool_cols].stack().dropna().reset_index().rename(columns={'level_1': 'Tool', 0: 'Date'})
                monthly_activity['Month'] = pd.to_datetime(monthly_activity['Date']).dt.to_period('M')
                avg_tools_per_month = monthly_activity.groupby('Month')['Tool'].nunique().mean() if not monthly_activity.empty else 0
                
                trend = "N/A"
                if len(activity_dates) > 1:
                    trend = "Stable"
                    timeline_midpoint = first_activity + (last_activity - first_activity) / 2
                    first_half_activity = monthly_activity[monthly_activity['Date'] <= timeline_midpoint]
                    second_half_activity = monthly_activity[monthly_activity['Date'] > timeline_midpoint]
                    if second_half_activity['Tool'].nunique() > first_half_activity['Tool'].nunique():
                        trend = "Increasing"
                    elif second_half_activity['Tool'].nunique() < first_half_activity['Tool'].nunique():
                        trend = "Decreasing"
                        
            consistency = (active_months / total_months_in_period) * 100 if total_months_in_period > 0 else 0
            
            user_metrics.append({
                'Email': email,
                'Usage Consistency (%)': consistency,
                'Overall Recency': last_activity,
                'Usage Complexity': complexity,
                'Avg Tools / Report': avg_tools_per_month,
                'Usage Trend': trend,
                'Appearances': user_data['Report Refresh Date'].nunique(),
                'First Appearance': first_activity
            })
            
        self.utilized_metrics_df = pd.DataFrame(user_metrics)
        
        # Calculate engagement scores
        if not self.utilized_metrics_df.empty:
            max_consistency = self.utilized_metrics_df['Usage Consistency (%)'].max()
            max_complexity = self.utilized_metrics_df['Usage Complexity'].max()
            max_avg_complexity = self.utilized_metrics_df['Avg Tools / Report'].max()
            
            self.utilized_metrics_df['consistency_norm'] = self.utilized_metrics_df['Usage Consistency (%)'] / max_consistency if max_consistency > 0 else 0
            self.utilized_metrics_df['complexity_norm'] = self.utilized_metrics_df['Usage Complexity'] / max_complexity if max_complexity > 0 else 0
            self.utilized_metrics_df['avg_complexity_norm'] = self.utilized_metrics_df['Avg Tools / Report'] / max_avg_complexity if max_avg_complexity > 0 else 0
            
            self.utilized_metrics_df['Engagement Score'] = (
                self.utilized_metrics_df['consistency_norm'] + 
                self.utilized_metrics_df['complexity_norm'] + 
                self.utilized_metrics_df['avg_complexity_norm']
            )
            
        return self.classify_users(usage_df, total_months_in_period)
        
    def classify_users(self, usage_df, total_months_in_period):
        """Classify users into categories"""
        reference_date = usage_df['Report Refresh Date'].max()
        thirty_days_ago = reference_date - timedelta(days=30)
        sixty_days_ago = reference_date - timedelta(days=60)
        ninety_days_ago = reference_date - timedelta(days=90)
        
        reallocation_emails, under_utilized_emails = set(), set()
        
        for _, row in self.utilized_metrics_df.iterrows():
            email = row['Email']
            is_new_user = pd.notna(row['First Appearance']) and row['First Appearance'] > thirty_days_ago
            
            if is_new_user:
                under_utilized_emails.add(email)
                continue
                
            if (row['Usage Complexity'] == 0) or \
               (pd.notna(row['Overall Recency']) and row['Overall Recency'] < ninety_days_ago) or \
               ((row['Usage Consistency (%)'] < 25) and not is_new_user):
                reallocation_emails.add(email)
                
            if (pd.notna(row['Overall Recency']) and (ninety_days_ago <= row['Overall Recency'] < sixty_days_ago)) or \
               (row['Usage Trend'] == 'Decreasing') or \
               (row['Appearances'] == 1) or \
               ((row['Usage Consistency (%)'] < 50) and not is_new_user):
                under_utilized_emails.add(email)
                
        # Create classification dataframes
        reallocation_df = self.utilized_metrics_df[self.utilized_metrics_df['Email'].isin(reallocation_emails)].copy()
        reallocation_df['Classification'] = 'For Reallocation'
        
        under_utilized_df = self.utilized_metrics_df[
            self.utilized_metrics_df['Email'].isin(under_utilized_emails) & 
            ~self.utilized_metrics_df['Email'].isin(reallocation_emails)
        ].copy()
        under_utilized_df['Classification'] = 'Under-Utilized'
        
        top_utilizer_emails = set(self.utilized_metrics_df['Email']) - reallocation_emails - under_utilized_emails
        top_utilizers_df = self.utilized_metrics_df[self.utilized_metrics_df['Email'].isin(top_utilizer_emails)].copy()
        top_utilizers_df['Classification'] = 'Top Utilizer'
        
        # Add justifications
        def get_justification(row, total_months):
            reasons = []
            is_new_user = pd.notna(row['First Appearance']) and row['First Appearance'] > thirty_days_ago
            if is_new_user:
                reasons.append("New user (in 30-day grace period)")
                
            if row['Usage Complexity'] == 0:
                reasons.append("No tool usage recorded")
            elif pd.notna(row['Overall Recency']) and row['Overall Recency'] < ninety_days_ago:
                reasons.append("No activity in 90+ days")
            elif pd.notna(row['Overall Recency']) and (ninety_days_ago <= row['Overall Recency'] < sixty_days_ago):
                reasons.append("No activity in 60-89 days")
                
            if row['Usage Trend'] == 'Decreasing':
                reasons.append("Downward usage trend")
                
            active_months = int(row['Usage Consistency (%)'] * total_months / 100)
            if row['Appearances'] == 1 and not is_new_user:
                reasons.append("Single report appearance")
            elif row['Usage Consistency (%)'] < 50 and not is_new_user:
                reasons.append(f"Low consistency (active in {active_months} of {total_months} months)")
                
            return "; ".join(reasons) if reasons else "High Engagement"
            
        top_utilizers_df['Justification'] = "High Engagement"
        under_utilized_df['Justification'] = [get_justification(row, total_months_in_period) for _, row in under_utilized_df.iterrows()]
        reallocation_df['Justification'] = [get_justification(row, total_months_in_period) for _, row in reallocation_df.iterrows()]
        
        # Sort dataframes
        top_utilizers_df.sort_values(by=['Engagement Score', 'Overall Recency'], ascending=[False, False], inplace=True)
        under_utilized_df.sort_values(by=['Engagement Score', 'Overall Recency'], ascending=[True, True], inplace=True)
        reallocation_df.sort_values(by=['Engagement Score', 'Overall Recency'], ascending=[True, True], inplace=True)
        
        return top_utilizers_df, under_utilized_df, reallocation_df
        
    def style_excel_sheet(self, worksheet, df):
        """Applies styling to an Excel worksheet."""
        # If the dataframe is empty, there is nothing to style.
        if df.empty:
            return

        # Header styling with dark background and white text
        header_fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_num, column_title in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        # Zebra striping for alternating rows
        stripe_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        for row_index in range(2, len(df) + 2):
            if row_index % 2 == 1:  # Apply to odd data rows (3, 5, 7...)
                for col_index in range(1, len(df.columns) + 1):
                    worksheet.cell(row=row_index, column=col_index).fill = stripe_fill

        # Auto-adjust column widths
        for col_num, column_title in enumerate(df.columns, 1):
            max_length = 0
            column_letter = get_column_letter(col_num)
            if len(str(column_title)) > max_length:
                max_length = len(str(column_title))
            for i, cell_value in enumerate(df[column_title], 2):
                if len(str(cell_value)) > max_length:
                    max_length = len(str(cell_value))
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Color scale and data bar formatting
        red_color, yellow_color, green_color = "F8696B", "FFEB84", "63BE7B"
        
        # Conditional formatting with color scales for Engagement Score
        if 'Engagement Score' in df.columns:
            score_col_letter = get_column_letter(df.columns.get_loc('Engagement Score') + 1)
            score_range = f"{score_col_letter}2:{score_col_letter}{len(df)+1}"
            worksheet.conditional_formatting.add(score_range,
                ColorScaleRule(start_type='min', start_color=red_color,
                               mid_type='percentile', mid_value=50, mid_color=yellow_color,
                               end_type='max', end_color=green_color))

        # Data bars for Usage Consistency percentage
        if 'Usage Consistency (%)' in df.columns:
            consistency_col_letter = get_column_letter(df.columns.get_loc('Usage Consistency (%)') + 1)
            consistency_range = f"{consistency_col_letter}2:{consistency_col_letter}{len(df)+1}"
            worksheet.conditional_formatting.add(consistency_range,
                DataBarRule(start_type='min', end_type='max', color=green_color))

    def create_visualizations(self, utilized_df, top_df, under_df, matched_df, output_folder):
        """Create visualizations for the Excel report"""
        charts = {}
        plt.style.use('default')
        
        try:
            # Engagement Score Distribution
            if 'Engagement Score' in utilized_df.columns and not utilized_df.empty:
                plt.figure(figsize=(10, 6))
                utilized_df['Engagement Score'].plot(kind='hist', bins=20, title='Distribution of User Engagement Score')
                plt.xlabel('Engagement Score (Consistency + Complexity + Avg Tools/Rpt)')
                plt.ylabel('Number of Users')
                plt.tight_layout()
                charts['engagement_score_hist'] = os.path.join(output_folder, 'engagement_score_hist.png')
                plt.savefig(charts['engagement_score_hist'], dpi=150, bbox_inches='tight')
                plt.close()
            
            # Tool Usage by Top Utilizers
            if not matched_df.empty and not top_df.empty:
                tool_cols = [col for col in matched_df.columns if 'Last activity date of' in col]
                top_user_activity = matched_df[matched_df['User Principal Name'].isin(top_df['Email'])]
                if not top_user_activity.empty and tool_cols:
                    tool_usage_counts = top_user_activity[tool_cols].notna().sum().sort_values(ascending=False)
                    if not tool_usage_counts.empty:
                        tool_usage_counts.index = tool_usage_counts.index.str.replace('Last activity date of ', '').str.replace(r' \(UTC\)', '')
                        plt.figure(figsize=(12, 7))
                        tool_usage_counts.plot(kind='bar', title='Most Commonly Used Tools by Top Utilizers')
                        plt.ylabel('Number of Top Users Using Tool')
                        plt.xticks(rotation=45, ha='right')
                        plt.tight_layout()
                        charts['top_utilizer_tools'] = os.path.join(output_folder, 'top_utilizer_tools.png')
                        plt.savefig(charts['top_utilizer_tools'], dpi=150, bbox_inches='tight')
                        plt.close()
            
            # Average Engagement Score Over Time
            if not matched_df.empty and 'Engagement Score' in utilized_df.columns:
                plot_data = pd.merge(
                    matched_df,
                    utilized_df[['Email', 'Engagement Score']],
                    left_on='User Principal Name',
                    right_on='Email',
                    how='left'
                )
                if not plot_data.empty and 'Engagement Score' in plot_data.columns:
                    trend_data = plot_data.groupby(pd.to_datetime(plot_data['Report Refresh Date']))['Engagement Score'].mean()
                    
                    if not trend_data.empty:
                        plt.figure(figsize=(12, 6))
                        trend_data.plot(kind='line', marker='o', linestyle='-', title='Average Engagement Score Over Time')
                        plt.ylabel('Average Engagement Score')
                        plt.xlabel('Report Date')
                        plt.grid(True)
                        plt.tight_layout()
                        charts['avg_engagement_trend'] = os.path.join(output_folder, 'avg_engagement_trend.png')
                        plt.savefig(charts['avg_engagement_trend'], dpi=150, bbox_inches='tight')
                        plt.close()
                        
            self.log("Visualizations created.")
            return charts
        except Exception as e:
            self.log(f"Error creating visualizations: {e}")
            return {}

    def create_excel_report(self, filename, top_utilizers_df, under_utilized_df, reallocation_df):
        """Create Excel report with full formatting"""
        try:
            # Create visualizations
            charts = self.create_visualizations(
                self.utilized_metrics_df, 
                top_utilizers_df, 
                under_utilized_df, 
                self.full_usage_data, 
                self.output_folder_path
            )
            
            with pd.ExcelWriter(filename, engine='openpyxl') as writer:
                # Define columns to include in sheets
                cols = ['Email', 'Classification', 'Usage Consistency (%)', 'Overall Recency', 
                       'Usage Complexity', 'Avg Tools / Report', 'Usage Trend', 'Engagement Score', 'Justification']
                
                # Create leaderboard data combining all classifications
                leaderboard_data = pd.concat([top_utilizers_df, under_utilized_df, reallocation_df]).sort_values(by="Engagement Score", ascending=False)
                
                # Prepare sheets with proper formatting
                sheets_to_process = {
                    'Leaderboard': leaderboard_data,
                    'Top Utilizers': top_utilizers_df,
                    'Under-Utilized': under_utilized_df,
                    'For Reallocation': reallocation_df
                }
                
                # Process each sheet
                for sheet_name, df in sheets_to_process.items():
                    if not df.empty:
                        df_to_write = df[cols].copy()
                        # Add rank column
                        df_to_write.insert(0, 'Rank', range(1, 1 + len(df_to_write)))
                        df_to_write.to_excel(writer, sheet_name=sheet_name, index=False, float_format="%.2f")
                        # Apply styling
                        self.style_excel_sheet(writer.sheets[sheet_name], df_to_write)

                # Create Summary & Visualizations sheet with embedded charts
                summary_ws = writer.book.create_sheet('Summary & Visualizations')
                try:
                    from openpyxl.drawing.image import Image as OpenpyxlImage
                    
                    # Add summary statistics as text
                    summary_ws['A1'] = 'Copilot License Evaluation Summary'
                    summary_ws['A1'].font = Font(bold=True, size=16)
                    summary_ws['A3'] = f'Total Users Analyzed: {len(self.utilized_metrics_df)}'
                    summary_ws['A4'] = f'Top Utilizers: {len(top_utilizers_df)}'
                    summary_ws['A5'] = f'Under-Utilized: {len(under_utilized_df)}'
                    summary_ws['A6'] = f'For Reallocation: {len(reallocation_df)}'
                    summary_ws['A7'] = f'Report Generated: {datetime.now().strftime("%B %d, %Y at %I:%M %p")}'
                    
                    # Embed charts if they exist
                    if 'engagement_score_hist' in charts and os.path.exists(charts['engagement_score_hist']):
                        img1 = OpenpyxlImage(charts['engagement_score_hist'])
                        img1.width = 600
                        img1.height = 360
                        summary_ws.add_image(img1, 'A10')
                    
                    if 'top_utilizer_tools' in charts and os.path.exists(charts['top_utilizer_tools']):
                        img2 = OpenpyxlImage(charts['top_utilizer_tools'])
                        img2.width = 700
                        img2.height = 420
                        summary_ws.add_image(img2, 'A35')
                    
                    if 'avg_engagement_trend' in charts and os.path.exists(charts['avg_engagement_trend']):
                        img3 = OpenpyxlImage(charts['avg_engagement_trend'])
                        img3.width = 700
                        img3.height = 360
                        summary_ws.add_image(img3, 'L10')
                        
                except Exception as e:
                    self.log(f"Error adding images to Excel: {e}")
                    
            self.log(f"Excel report created: {filename}")
            return True
        except Exception as e:
            self.log(f"Error creating Excel report: {e}")
            return False
            
    def create_leaderboard_html(self, filename):
        """Create HTML leaderboard with original Haleon theme styling"""
        try:
            if self.utilized_metrics_df is None or self.utilized_metrics_df.empty:
                self.log("No data available to generate leaderboard.")
                return False

            leaderboard_data = self.utilized_metrics_df.sort_values(by="Engagement Score", ascending=False)
            
            html_head = """
            <!DOCTYPE html>
            <html lang="en">
            <head>
                <meta charset="UTF-8">
                <meta name="viewport" content="width=device-width, initial-scale=1.0">
                <title>Leaderboard - Haleon Theme</title>
                <script src="https://cdn.tailwindcss.com"></script>
                <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;600;700&display=swap" rel="stylesheet">
                <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
                <style>
                    body { 
                        font-family: 'Inter', sans-serif; 
                        background-color: #f3f4f6;
                    }
                    .leaderboard-component {
                        border-radius: 1rem; 
                        box-shadow: 0 10px 15px -3px rgb(0 0 0 / 0.1), 0 4px 6px -4px rgb(0 0 0 / 0.1);
                        overflow: hidden;
                        border: 1px solid #e5e7eb;
                    }
                    .title-banner {
                        background-color: #000000;
                    }
                    .table-container {
                        background-color: #FFFFFF;
                    }
                    .table-header { 
                        background-color: #2d3748;
                    }
                    .table-row:nth-child(even) { background-color: #f9fafb; }
                    .table-row:hover { 
                        background-color: #f0f0f0; 
                    }
                    .rank-badge { 
                        font-weight: 700; 
                        width: 2.5rem; 
                        height: 2.5rem; 
                        display: flex; 
                        align-items: center; 
                        justify-content: center; 
                        border-radius: 50%;
                        color: #000;
                    }
                    .progress-bar-container { 
                        background-color: #e5e7eb; 
                        border-radius: 9999px; 
                        height: 8px; 
                        width: 100%; 
                    }
                    .progress-bar { 
                        background: #39FF14; 
                        border-radius: 9999px; 
                        height: 100%; 
                    }
                    .trend-icon.Increasing { color: #16a34a; }
                    .trend-icon.Stable { color: #f59e0b; }
                    .trend-icon.Decreasing { color: #ef4444; }
                    .trend-icon.N\\/A { color: #6b7280; }
                    .neon-green-text {
                        color: #16a34a;
                    }
                    .user-email {
                        font-weight: 600;
                        color: #000000;
                    }
                </style>
            </head>
            <body class="p-4 sm:p-6 lg:p-8">
                <div class="leaderboard-component w-full max-w-5xl mx-auto">
                    <div class="title-banner p-6 text-center">
                        <h1 class="text-4xl font-bold text-white mb-2">Copilot Usage Leaderboard</h1>
                        <p class="text-gray-300">Ranking by Engagement Score</p>
                    </div>
                    <div class="table-container">
                        <div class="overflow-x-auto">
                            <div class="min-w-full inline-block align-middle">
                                <div class="table-header">
                                    <div class="grid grid-cols-12 gap-4 px-6 py-4 text-left text-xs font-bold uppercase tracking-wider text-white">
                                        <div class="col-span-1">Rank</div>
                                        <div class="col-span-5">User</div>
                                        <div class="col-span-2 text-center">Consistency</div>
                                        <div class="col-span-2 text-center">Trend</div>
                                        <div class="col-span-2 text-right">Engagement</div>
                                    </div>
                                </div>
                                <div class="divide-y divide-gray-200">
            """
            
            html_rows = ""
            max_score = leaderboard_data['Engagement Score'].max() if not leaderboard_data.empty else 3.0
            
            for rank, (_, user_row) in enumerate(leaderboard_data.iterrows(), 1):
                try:
                    if pd.isna(user_row['Email']) or not isinstance(user_row['Email'], str): 
                        continue

                    score_percentage = (user_row['Engagement Score'] / max_score) * 100 if max_score > 0 else 0
                    hue = score_percentage * 1.2  # 0% -> 0 (red), 100% -> 120 (green)
                    badge_color = f"hsl({hue}, 80%, 50%)"
                    
                    trend = user_row['Usage Trend']
                    trend_icon_map = {
                        'Increasing': 'fa-arrow-trend-up', 
                        'Decreasing': 'fa-arrow-trend-down', 
                        'Stable': 'fa-minus', 
                        'N/A': 'fa-question'
                    }
                    trend_icon = f"fa-solid {trend_icon_map.get(trend, 'fa-minus')}"
                    
                    html_rows += f"""
                                    <div class="grid grid-cols-12 gap-4 px-6 py-3 items-center table-row text-gray-800">
                                        <div class="col-span-1"><div class="rank-badge" style="background-color: {badge_color};"><span>{rank}</span></div></div>
                                        <div class="col-span-5"><div class="user-email">{user_row['Email']}</div></div>
                                        <div class="col-span-2 text-center"><div class="text-sm font-semibold neon-green-text">{user_row['Usage Consistency (%)']:.1f}%</div><div class="progress-bar-container mt-1"><div class="progress-bar" style="width: {user_row['Usage Consistency (%)']}%"></div></div></div>
                                        <div class="col-span-2 text-center"><i class="trend-icon {trend} {trend_icon} fa-lg"></i></div>
                                        <div class="col-span-2 text-right"><div class="text-sm font-bold neon-green-text">{user_row['Engagement Score']:.2f}</div></div>
                                    </div>
                    """
                except Exception as e:
                    self.log(f"Error processing leaderboard row for {user_row.get('Email', 'N/A')}: {e}")
            
            html_foot = """
                                </div>
                            </div>
                        </div>
                    </div>
                </div>
            </body>
            </html>
            """
            
            full_html = html_head + html_rows + html_foot
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(full_html)
            self.log(f"HTML Leaderboard created: {filename}")
            return True
        except Exception as e:
            self.log(f"Error creating HTML leaderboard: {e}")
            return False
            
    def get_filter_options(self):
        """Get available filter options from target user data"""
        if self.target_user_data is None:
            return {}
            
        try:
            options = {
                'companies': sorted(self.target_user_data['Company'].dropna().unique().tolist()),
                'departments': sorted(self.target_user_data['Department'].dropna().unique().tolist()),
                'locations': sorted(self.target_user_data['City'].dropna().unique().tolist()),
                'managers': []
            }
            
            # Extract managers from manager lines
            all_managers = set()
            for chain in self.target_user_data['ManagerLine'].dropna():
                managers_in_chain = [m.strip() for m in chain.split('->')]
                all_managers.update(managers_in_chain)
            options['managers'] = sorted(list(all_managers))
            
            return options
        except Exception as e:
            self.log(f"Error getting filter options: {e}")
            return {}
            
def main():
    parser = argparse.ArgumentParser(description='Copilot Usage Analyzer')
    parser.add_argument('--target-users', help='Path to target users CSV file')
    parser.add_argument('--usage-reports', nargs='+', required=True, help='Paths to usage report files')
    parser.add_argument('--output-dir', required=True, help='Output directory for reports')
    parser.add_argument('--filters', help='JSON string with filter options')
    
    args = parser.parse_args()
    
    try:
        analyzer = CopilotAnalyzer()
        analyzer.output_folder_path = args.output_dir
        
        # Ensure output directory exists
        os.makedirs(args.output_dir, exist_ok=True)
        
        # Load target users if provided
        filtered_target_df = None
        if args.target_users:
            if not analyzer.load_target_users(args.target_users):
                sys.exit(1)
                
            # Apply filters if provided
            if args.filters:
                filters = json.loads(args.filters)
                filtered_target_df = analyzer.apply_filters(filters)
                
        # Load usage reports
        if not analyzer.load_usage_reports(args.usage_reports):
            sys.exit(1)
            
        # Perform analysis
        analyzer.log("Starting analysis...")
        top_utilizers_df, under_utilized_df, reallocation_df = analyzer.analyze_users(filtered_target_df)
        
        # Generate reports
        today_str = datetime.now().strftime("%d-%B-%Y")
        excel_filename = os.path.join(args.output_dir, f"{today_str}_Copilot_License_Evaluation.xlsx")
        html_filename = os.path.join(args.output_dir, "leaderboard.html")
        
        analyzer.create_excel_report(excel_filename, top_utilizers_df, under_utilized_df, reallocation_df)
        analyzer.create_leaderboard_html(html_filename)
        
        # Prepare detailed user data for web interface
        detailed_users = []
        for _, row in analyzer.utilized_metrics_df.iterrows():
            user_data = {
                'email': row['Email'],
                'engagementScore': float(row['Engagement Score']),
                'consistencyPercent': float(row['Usage Consistency (%)']),
                'complexityScore': float(row['Usage Complexity']),
                'avgToolsPerReport': float(row['Avg Tools / Report']),
                'trend': row['Usage Trend'],
                'appearances': int(row['Appearances']),
                'firstAppearance': row['First Appearance'].isoformat() if pd.notna(row['First Appearance']) else None,
                'lastActivity': row['Overall Recency'].isoformat() if pd.notna(row['Overall Recency']) else None,
                'classification': None,  # Will be set below
                'justification': None,  # Will be set below
                'riskLevel': 'Low'  # Default, will be updated below
            }
            detailed_users.append(user_data)
        
        # Add classification and justification data
        for _, row in top_utilizers_df.iterrows():
            user_data = next((u for u in detailed_users if u['email'] == row['Email']), None)
            if user_data:
                user_data['classification'] = 'Top Utilizer'
                user_data['justification'] = row['Justification']
                user_data['riskLevel'] = 'Low'
        
        for _, row in under_utilized_df.iterrows():
            user_data = next((u for u in detailed_users if u['email'] == row['Email']), None)
            if user_data:
                user_data['classification'] = 'Under-Utilized'
                user_data['justification'] = row['Justification']
                user_data['riskLevel'] = 'Medium'
        
        for _, row in reallocation_df.iterrows():
            user_data = next((u for u in detailed_users if u['email'] == row['Email']), None)
            if user_data:
                user_data['classification'] = 'For Reallocation'
                user_data['justification'] = row['Justification']
                user_data['riskLevel'] = 'High'
        
        # Generate tool usage data from usage reports
        tool_usage_data = {}
        if analyzer.full_usage_data is not None:
            tool_cols = [col for col in analyzer.full_usage_data.columns if 'Last activity date of' in col]
            for email in [u['email'] for u in detailed_users]:
                user_usage = analyzer.full_usage_data[analyzer.full_usage_data['User Principal Name'] == email]
                tools_used = []
                for col in tool_cols:
                    if user_usage[col].notna().any():
                        tool_name = col.replace('Last activity date of ', '').replace(' (UTC)', '')
                        tools_used.append(tool_name)
                tool_usage_data[email] = tools_used
        
        # Add tool usage to detailed users
        for user_data in detailed_users:
            user_data['toolsUsed'] = tool_usage_data.get(user_data['email'], [])
            
            # Generate mock monthly activity data for visualization
            user_data['monthlyActivity'] = []
            months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            for month in months:
                if user_data['classification'] == 'Top Utilizer':
                    activity = {'month': month, 'toolsUsed': 3 + (hash(month + user_data['email']) % 3), 'complexity': 6 + (hash(month + user_data['email']) % 4)}
                elif user_data['classification'] == 'Under-Utilized':
                    activity = {'month': month, 'toolsUsed': (hash(month + user_data['email']) % 2), 'complexity': (hash(month + user_data['email']) % 3)}
                else:  # For Reallocation
                    activity = {'month': month, 'toolsUsed': 0, 'complexity': 0}
                user_data['monthlyActivity'].append(activity)
            
            # Generate report dates
            user_data['reportDates'] = []
            for i in range(user_data['appearances']):
                user_data['reportDates'].append(f"2024-{(i % 12) + 1:02d}-01T00:00:00.000Z")

        # Output results as JSON for web interface
        results = {
            'status': 'success',
            'summary': {
                'total_users': len(analyzer.utilized_metrics_df),
                'top_utilizers': len(top_utilizers_df),
                'under_utilized': len(under_utilized_df),
                'for_reallocation': len(reallocation_df)
            },
            'files': {
                'excel': excel_filename,
                'html': html_filename
            },
            'detailed_users': detailed_users
        }
        
        print(json.dumps(results))
        
    except Exception as e:
        error_result = {
            'status': 'error',
            'message': str(e)
        }
        print(json.dumps(error_result))
        sys.exit(1)
        
if __name__ == "__main__":
    main()
